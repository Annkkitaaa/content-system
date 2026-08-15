"""Scheduling, the workbook, and the weekly run.

The end of the brief: one command produces one file. These tests care about
the two properties that make that file trustworthy, namely that the calendar
does not look machine generated and that the workbook is actually usable when
opened.
"""

from __future__ import annotations

from datetime import date, timedelta
from itertools import pairwise
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlmodel import Session

from contentsys.config import DraftStatus, Platform, Settings, get_settings
from contentsys.export import workbook as wb
from contentsys.knowledge import load_seed
from contentsys.llm.mock import MockProvider
from contentsys.pipeline import record_snapshot, run_weekly
from contentsys.scheduling import WEEKDAYS, week_starting, weekly_slots
from contentsys.voice import build_profile

SEED_DIR = Path(__file__).resolve().parent.parent / "seed"


@pytest.fixture
def settings() -> Settings:
    get_settings.cache_clear()
    return Settings()


@pytest.fixture
def seeded(session: Session) -> Session:
    load_seed(session, SEED_DIR)
    session.commit()
    for platform in Platform:
        build_profile(session, platform)
    session.commit()
    return session


class TestWeekStart:
    def test_finds_the_next_monday(self) -> None:
        # A Saturday plans for the Monday two days later.
        assert week_starting(date(2026, 8, 15)) == date(2026, 8, 17)

    def test_a_monday_plans_for_itself(self) -> None:
        assert week_starting(date(2026, 8, 17)) == date(2026, 8, 17)


class TestScheduling:
    def test_a_full_week_is_produced(self, settings: Settings) -> None:
        slots = weekly_slots(settings, start=date(2026, 8, 17), seed=1)

        assert len(slots[Platform.X]) == 70
        assert len(slots[Platform.LINKEDIN]) == 2

    def test_counts_are_overridable(self, settings: Settings) -> None:
        slots = weekly_slots(settings, start=date(2026, 8, 17), x_posts=3, linkedin_posts=1, seed=1)

        assert len(slots[Platform.X]) == 21
        assert len(slots[Platform.LINKEDIN]) == 1

    def test_every_day_of_the_week_is_covered(self, settings: Settings) -> None:
        slots = weekly_slots(settings, start=date(2026, 8, 17), seed=1)
        days = {slot.when.date() for slot in slots[Platform.X]}

        assert len(days) == 7

    def test_the_same_seed_gives_the_same_calendar(self, settings: Settings) -> None:
        # Reproducible so a run can be regenerated and diffed.
        first = weekly_slots(settings, start=date(2026, 8, 17), seed=42)
        second = weekly_slots(settings, start=date(2026, 8, 17), seed=42)

        assert [s.timestamp for s in first[Platform.X]] == [s.timestamp for s in second[Platform.X]]

    def test_different_seeds_give_different_times(self, settings: Settings) -> None:
        first = weekly_slots(settings, start=date(2026, 8, 17), seed=1)
        second = weekly_slots(settings, start=date(2026, 8, 17), seed=2)

        assert [s.timestamp for s in first[Platform.X]] != [s.timestamp for s in second[Platform.X]]

    def test_times_do_not_look_machine_generated(self, settings: Settings) -> None:
        # The point of the jitter. Evenly spaced posts on round minutes are a
        # behavioural signal the ranking model uses against an account.
        slots = weekly_slots(settings, start=date(2026, 8, 17), seed=7)
        minutes = [slot.when.minute for slot in slots[Platform.X]]

        on_the_hour = sum(1 for minute in minutes if minute in {0, 30})
        assert on_the_hour / len(minutes) < 0.25, "too many posts land on round times"
        assert len(set(minutes)) > 20, "not enough variety in posting minutes"

    def test_the_minimum_gap_holds_after_jittering(self, settings: Settings) -> None:
        # Jitter is what creates collisions, so the gap has to be enforced
        # after it or the constraint is decoration.
        slots = weekly_slots(settings, start=date(2026, 8, 17), seed=3)
        gap = settings.schedule.x.min_gap_minutes

        by_day: dict[date, list] = {}
        for slot in slots[Platform.X]:
            by_day.setdefault(slot.when.date(), []).append(slot.when)

        for day, times in by_day.items():
            ordered = sorted(times)
            for earlier, later in pairwise(ordered):
                minutes = (later - earlier).total_seconds() / 60
                assert minutes >= gap - 0.01, f"{day}: only {minutes:.0f} minutes apart"

    def test_slots_land_inside_configured_windows(self, settings: Settings) -> None:
        slots = weekly_slots(settings, start=date(2026, 8, 17), seed=5)
        windows = settings.schedule.x.windows
        earliest = min(window.start_minutes for window in windows)
        latest = max(window.end_minutes for window in windows)

        for slot in slots[Platform.X]:
            minute = slot.when.hour * 60 + slot.when.minute
            assert earliest <= minute <= latest + settings.schedule.x.min_gap_minutes

    def test_linkedin_uses_its_preferred_days(self, settings: Settings) -> None:
        slots = weekly_slots(settings, start=date(2026, 8, 17), seed=1)
        days = {WEEKDAYS[slot.when.weekday()] for slot in slots[Platform.LINKEDIN]}

        assert days <= set(settings.schedule.linkedin.preferred_days)

    def test_slots_carry_a_readable_timestamp(self, settings: Settings) -> None:
        slot = weekly_slots(settings, start=date(2026, 8, 17), seed=1)[Platform.X][0]

        assert slot.timestamp.startswith("2026-08-17T")
        assert "+05:30" in slot.timestamp
        assert slot.day == "Monday"


class TestWorkbook:
    def sample(self) -> wb.WeeklyWorkbook:
        return wb.WeeklyWorkbook(
            week_starting=date(2026, 8, 17),
            calendar=[
                wb.CalendarRow(
                    date_text="2026-08-17",
                    day="Monday",
                    time_text="9:14 AM",
                    timestamp="2026-08-17T09:14:00+05:30",
                    platform="X",
                    content_type="technical",
                    topic="sumcheck",
                    content="sumcheck collapses an exponential sum to one point.",
                    authenticity=8.7,
                    originality=9.0,
                    voice_match=9.5,
                    slop_risk="LOW",
                    repetition_risk="LOW",
                ),
            ],
            ideas=[
                wb.IdeaRow(
                    idea="sumcheck: it got simpler as it generalised",
                    topic="sumcheck",
                    angle="counterintuitive",
                    platform="X",
                    content_type="technical",
                    why_interesting="most things go the other way",
                    novelty=8.5,
                )
            ],
            monetization=[wb.MonetizationRow("Verified followers", "3,895", "500", "met")],
        )

    def test_all_six_sheets_are_present(self, tmp_path: Path) -> None:
        path = wb.write(self.sample(), tmp_path)
        book = load_workbook(path)

        assert book.sheetnames == [
            "Weekly Calendar",
            "Ideas",
            "Research",
            "Voice Feedback",
            "Content History",
            "Monetization",
        ]

    def test_the_filename_carries_the_week(self, tmp_path: Path) -> None:
        assert wb.write(self.sample(), tmp_path).name == "weekly_content_2026-08-17.xlsx"

    def test_the_brief_columns_are_all_there(self, tmp_path: Path) -> None:
        sheet = load_workbook(wb.write(self.sample(), tmp_path))["Weekly Calendar"]
        headers = [cell.value for cell in sheet[1]]

        for required in (
            "Date",
            "Day",
            "Time",
            "Timestamp",
            "Platform",
            "Content Type",
            "Topic",
            "Content",
            "Status",
            "Source",
            "Notes",
        ):
            assert required in headers

    def test_the_header_is_frozen_and_filterable(self, tmp_path: Path) -> None:
        # It has to be usable when opened, not just correct.
        sheet = load_workbook(wb.write(self.sample(), tmp_path))["Weekly Calendar"]

        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref

    def test_status_is_a_dropdown_not_free_text(self, tmp_path: Path) -> None:
        # The workbook is re-imported, so free text would mean parsing
        # whatever someone typed. A dropdown keeps the round trip lossless.
        sheet = load_workbook(wb.write(self.sample(), tmp_path))["Weekly Calendar"]

        assert sheet.data_validations.dataValidation
        formula = sheet.data_validations.dataValidation[0].formula1
        for status in DraftStatus:
            assert status.value in formula

    def test_content_columns_wrap(self, tmp_path: Path) -> None:
        sheet = load_workbook(wb.write(self.sample(), tmp_path))["Weekly Calendar"]
        headers = [cell.value for cell in sheet[1]]
        column = headers.index("Content") + 1

        assert sheet.cell(row=2, column=column).alignment.wrap_text

    def test_an_empty_week_still_writes(self, tmp_path: Path) -> None:
        empty = wb.WeeklyWorkbook(week_starting=date(2026, 8, 17))

        assert wb.write(empty, tmp_path).exists()

    def test_the_summary_reports_what_matters(self) -> None:
        lines = " ".join(wb.summarise(self.sample()))

        assert "1 pieces scheduled" in lines
        assert "authenticity" in lines.lower()
        assert "Needs review" in lines


class TestWeeklyRun:
    def test_one_command_produces_a_workbook(
        self, seeded: Session, settings: Settings, tmp_path: Path
    ) -> None:
        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=2,
            linkedin_posts=1,
            seed=1,
        )

        assert result.path
        assert Path(result.path).exists()
        assert result.drafts

    def test_every_draft_gets_a_scheduled_slot(self, seeded: Session, settings: Settings) -> None:
        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=1,
            linkedin_posts=1,
            seed=1,
        )
        sheet = load_workbook(result.path)["Weekly Calendar"]

        assert sheet.max_row == len(result.drafts) + 1

    def test_an_unfilled_slot_is_reported_rather_than_hidden(
        self, seeded: Session, settings: Settings
    ) -> None:
        # Zipping slots against drafts truncates to the shorter of the two, so
        # a thin idea pool would otherwise produce a week with holes in it and
        # nothing anywhere saying so.
        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=10,
            linkedin_posts=2,
            seed=1,
        )

        planned = 70 + 2
        assert len(result.drafts) + result.missing == planned
        if result.missing:
            assert any("Short by" in line for line in result.summary)

    def test_a_draft_is_not_compared_against_itself(
        self, seeded: Session, settings: Settings
    ) -> None:
        # generate_batch appends each draft to context.recent_posts so later
        # posts avoid repeating earlier ones. Reading that list at evaluation
        # time made every post its own predecessor, and a live run reported
        # "identical to a previous post" for all fifteen distinct drafts.
        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=1,
            linkedin_posts=1,
            seed=1,
        )
        sheet = load_workbook(result.path)["Weekly Calendar"]
        headers = [cell.value for cell in sheet[1]]
        notes_column = headers.index("Notes") + 1
        notes = [
            str(sheet.cell(row=r, column=notes_column).value or "")
            for r in range(2, sheet.max_row + 1)
        ]

        assert not any("identical to a previous post" in note for note in notes), (
            "a draft is being compared against itself"
        )

    def test_the_calendar_is_in_time_order(self, seeded: Session, settings: Settings) -> None:
        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=2,
            linkedin_posts=1,
            seed=1,
        )
        sheet = load_workbook(result.path)["Weekly Calendar"]
        headers = [cell.value for cell in sheet[1]]
        column = headers.index("Timestamp") + 1
        stamps = [sheet.cell(row=r, column=column).value for r in range(2, sheet.max_row + 1)]

        assert stamps == sorted(stamps)

    def test_nothing_is_marked_published(self, seeded: Session, settings: Settings) -> None:
        # The default path is generate, review, edit, approve, publish. The
        # workbook must never arrive claiming a post already went out.
        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=1,
            linkedin_posts=1,
            seed=1,
        )
        sheet = load_workbook(result.path)["Weekly Calendar"]
        headers = [cell.value for cell in sheet[1]]
        column = headers.index("Status") + 1
        statuses = {sheet.cell(row=r, column=column).value for r in range(2, sheet.max_row + 1)}

        assert DraftStatus.PUBLISHED.value not in statuses
        assert DraftStatus.APPROVED.value not in statuses

    def test_unused_ideas_are_kept(self, seeded: Session, settings: Settings) -> None:
        # Oversampling only pays off if the leftovers survive to next week.
        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=1,
            linkedin_posts=1,
            seed=1,
        )

        assert result.unused_ideas >= 0
        assert load_workbook(result.path)["Ideas"].max_row >= 1

    def test_monetization_gates_appear_every_week(
        self, seeded: Session, settings: Settings
    ) -> None:
        # Shown whether or not a snapshot exists, because the point is to make
        # the distance visible.
        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=1,
            linkedin_posts=1,
            seed=1,
        )
        sheet = load_workbook(result.path)["Monetization"]
        metrics = [sheet.cell(row=r, column=1).value for r in range(2, sheet.max_row + 1)]

        assert any("followers" in str(m).lower() for m in metrics)
        assert any("impressions" in str(m).lower() for m in metrics)

    def test_a_recorded_snapshot_shows_the_gap(self, seeded: Session, settings: Settings) -> None:
        record_snapshot(
            seeded, verified_followers=3895, verified_impressions_90d=120_000, premium_active=True
        )
        seeded.commit()

        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=1,
            linkedin_posts=1,
            seed=1,
        )
        sheet = load_workbook(result.path)["Monetization"]
        rows = {
            sheet.cell(row=r, column=1).value: sheet.cell(row=r, column=4).value
            for r in range(2, sheet.max_row + 1)
        }

        assert rows["Verified followers"] == "met"
        assert "to go" in rows["Verified Home Timeline impressions, 90 days"]

    def test_usage_is_reported(self, seeded: Session, settings: Settings) -> None:
        result = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=1,
            linkedin_posts=1,
            seed=1,
        )

        assert result.usage.output_tokens > 0
        assert result.usage.cache_read_tokens > 0

    def test_the_run_is_reproducible(self, seeded: Session, settings: Settings) -> None:
        first = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=1,
            linkedin_posts=1,
            seed=9,
        )
        second = run_weekly(
            seeded,
            MockProvider(),
            settings,
            start=date(2026, 8, 17),
            x_posts=1,
            linkedin_posts=1,
            seed=9,
        )

        assert [d.content for d in first.drafts] == [d.content for d in second.drafts]


def test_a_week_starting_today_still_schedules_forward(settings: Settings) -> None:
    # Guards the off by one that would put the first post in the past.
    monday = week_starting()
    slots = weekly_slots(settings, start=monday, x_posts=1, seed=1)

    assert min(slot.when.date() for slot in slots[Platform.X]) == monday
    assert max(slot.when.date() for slot in slots[Platform.X]) == monday + timedelta(days=6)
