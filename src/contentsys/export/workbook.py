"""The weekly Excel workbook.

This is the deliverable, not an export feature. The whole system exists to
produce one file each week that can be opened, read, edited and worked from.

Two things follow from that. It has to be readable without explanation: frozen
headers, sensible column widths, wrapped text, colour that means something.
And it has to be editable, which means the Status column is a dropdown rather
than free text, so what comes back in is the same vocabulary that went out.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from contentsys.config import DraftStatus, Platform

HEADER_FILL = PatternFill("solid", fgColor="16161A")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(size=11)
MUTED_FONT = Font(size=10, color="6B6B76")
THIN = Side(style="thin", color="D6D3CC")
BORDER = Border(bottom=THIN)

GOOD_FILL = PatternFill("solid", fgColor="DCF5E4")
WARN_FILL = PatternFill("solid", fgColor="FDF0D5")
BAD_FILL = PatternFill("solid", fgColor="FADCD9")


@dataclass
class CalendarRow:
    """One scheduled piece of content."""

    date_text: str
    day: str
    time_text: str
    timestamp: str
    platform: str
    content_type: str
    topic: str
    content: str
    status: str = DraftStatus.DRAFT.value
    authenticity: float | None = None
    originality: float | None = None
    slop_risk: str = ""
    technical_accuracy: float | None = None
    voice_match: float | None = None
    repetition_risk: str = ""
    diagram: str = ""
    source: str = ""
    notes: str = ""


@dataclass
class IdeaRow:
    idea: str
    topic: str
    angle: str
    platform: str
    content_type: str
    why_interesting: str
    personal_connection: str = ""
    novelty: float | None = None
    status: str = DraftStatus.IDEA.value


@dataclass
class ResearchRow:
    topic: str
    source: str
    url: str = ""
    key_fact: str = ""
    potential_angle: str = ""
    used_in: str = ""


@dataclass
class FeedbackRow:
    original: str
    edited: str
    what_changed: str
    preference_learned: str
    recorded_on: str


@dataclass
class HistoryRow:
    date_text: str
    platform: str
    content: str
    topic: str
    content_type: str
    published: str = "No"
    performance: str = ""


@dataclass
class MonetizationRow:
    metric: str
    current: str
    target: str
    status: str
    note: str = ""


@dataclass
class WeeklyWorkbook:
    """Everything one week's file contains."""

    week_starting: date
    calendar: list[CalendarRow] = field(default_factory=list)
    ideas: list[IdeaRow] = field(default_factory=list)
    research: list[ResearchRow] = field(default_factory=list)
    feedback: list[FeedbackRow] = field(default_factory=list)
    history: list[HistoryRow] = field(default_factory=list)
    monetization: list[MonetizationRow] = field(default_factory=list)

    def filename(self) -> str:
        return f"weekly_content_{self.week_starting.isoformat()}.xlsx"


def _write_sheet(
    sheet: Worksheet,
    headers: list[str],
    rows: list[list[Any]],
    widths: dict[str, int],
    *,
    wrap_columns: set[str] | None = None,
) -> None:
    """Write a sheet that a person can actually read.

    Frozen header, autofilter, per column widths, wrapping only where the
    content is prose. Wrapping everything makes a spreadsheet where nothing
    lines up.
    """
    sheet.append(headers)
    for index, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left")
    sheet.row_dimensions[1].height = 26

    wrap = wrap_columns or set()
    for row in rows:
        sheet.append(row)

    for row_index in range(2, sheet.max_row + 1):
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=header in wrap,
                horizontal="left",
            )

    for column_index, header in enumerate(headers, start=1):
        letter = get_column_letter(column_index)
        sheet.column_dimensions[letter].width = widths.get(header, 16)

    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{sheet.max_row}"


def _score_rules(sheet: Worksheet, headers: list[str], names: list[str]) -> None:
    """Colour the score columns so a weak week is visible at a glance."""
    if sheet.max_row < 2:
        return
    for name in names:
        if name not in headers:
            continue
        letter = get_column_letter(headers.index(name) + 1)
        span = f"{letter}2:{letter}{sheet.max_row}"
        sheet.conditional_formatting.add(
            span, CellIsRule(operator="greaterThanOrEqual", formula=["8"], fill=GOOD_FILL)
        )
        sheet.conditional_formatting.add(
            span, CellIsRule(operator="between", formula=["6", "7.999"], fill=WARN_FILL)
        )
        sheet.conditional_formatting.add(
            span, CellIsRule(operator="lessThan", formula=["6"], fill=BAD_FILL)
        )


def _status_dropdown(sheet: Worksheet, headers: list[str]) -> None:
    """Constrain Status to the known vocabulary.

    The workbook is re-imported, so free text here would mean parsing whatever
    someone typed. A dropdown keeps the round trip lossless.
    """
    if "Status" not in headers or sheet.max_row < 2:
        return
    letter = get_column_letter(headers.index("Status") + 1)
    options = ",".join(status.value for status in DraftStatus)
    validation = DataValidation(type="list", formula1=f'"{options}"', allow_blank=False)
    validation.error = "Pick one of the listed statuses."
    validation.errorTitle = "Unknown status"
    sheet.add_data_validation(validation)
    validation.add(f"{letter}2:{letter}{sheet.max_row + 200}")


CALENDAR_HEADERS = [
    "Date",
    "Day",
    "Time",
    "Timestamp",
    "Platform",
    "Content Type",
    "Topic",
    "Content",
    "Status",
    "Authenticity",
    "Originality",
    "Voice Match",
    "AI Slop Risk",
    "Repetition Risk",
    "Technical Accuracy",
    "Diagram",
    "Source",
    "Notes",
]

CALENDAR_WIDTHS = {
    "Date": 12,
    "Day": 11,
    "Time": 10,
    "Timestamp": 26,
    "Platform": 10,
    "Content Type": 18,
    "Topic": 26,
    "Content": 70,
    "Status": 12,
    "Authenticity": 12,
    "Originality": 11,
    "Voice Match": 12,
    "AI Slop Risk": 13,
    "Repetition Risk": 15,
    "Technical Accuracy": 17,
    "Diagram": 30,
    "Source": 22,
    "Notes": 48,
}


def build(workbook: WeeklyWorkbook) -> Workbook:
    """Assemble the six sheets."""
    book = Workbook()

    calendar = book.active
    calendar.title = "Weekly Calendar"
    _write_sheet(
        calendar,
        CALENDAR_HEADERS,
        [
            [
                row.date_text,
                row.day,
                row.time_text,
                row.timestamp,
                row.platform,
                row.content_type,
                row.topic,
                row.content,
                row.status,
                row.authenticity,
                row.originality,
                row.voice_match,
                row.slop_risk,
                row.repetition_risk,
                row.technical_accuracy,
                row.diagram,
                row.source,
                row.notes,
            ]
            for row in workbook.calendar
        ],
        CALENDAR_WIDTHS,
        wrap_columns={"Content", "Notes", "Topic"},
    )
    _score_rules(
        calendar,
        CALENDAR_HEADERS,
        ["Authenticity", "Originality", "Voice Match", "Technical Accuracy"],
    )
    _status_dropdown(calendar, CALENDAR_HEADERS)

    ideas_headers = [
        "Idea",
        "Topic",
        "Angle",
        "Platform",
        "Content Type",
        "Why Interesting",
        "Personal Connection",
        "Novelty Score",
        "Status",
    ]
    _write_sheet(
        book.create_sheet("Ideas"),
        ideas_headers,
        [
            [
                row.idea,
                row.topic,
                row.angle,
                row.platform,
                row.content_type,
                row.why_interesting,
                row.personal_connection,
                row.novelty,
                row.status,
            ]
            for row in workbook.ideas
        ],
        {"Idea": 40, "Angle": 50, "Why Interesting": 46, "Personal Connection": 30},
        wrap_columns={"Idea", "Angle", "Why Interesting", "Personal Connection"},
    )

    research_headers = ["Topic", "Source", "URL", "Key Fact", "Potential Angle", "Used In"]
    _write_sheet(
        book.create_sheet("Research"),
        research_headers,
        [
            [row.topic, row.source, row.url, row.key_fact, row.potential_angle, row.used_in]
            for row in workbook.research
        ],
        {"Source": 34, "URL": 40, "Key Fact": 50, "Potential Angle": 44},
        wrap_columns={"Key Fact", "Potential Angle"},
    )

    feedback_headers = [
        "Original Draft",
        "My Edited Version",
        "What Changed",
        "Voice Preference Learned",
        "Date",
    ]
    _write_sheet(
        book.create_sheet("Voice Feedback"),
        feedback_headers,
        [
            [row.original, row.edited, row.what_changed, row.preference_learned, row.recorded_on]
            for row in workbook.feedback
        ],
        {
            "Original Draft": 55,
            "My Edited Version": 55,
            "What Changed": 40,
            "Voice Preference Learned": 34,
        },
        wrap_columns={"Original Draft", "My Edited Version", "What Changed"},
    )

    history_headers = [
        "Date",
        "Platform",
        "Content",
        "Topic",
        "Content Type",
        "Published",
        "Performance",
    ]
    _write_sheet(
        book.create_sheet("Content History"),
        history_headers,
        [
            [
                row.date_text,
                row.platform,
                row.content,
                row.topic,
                row.content_type,
                row.published,
                row.performance,
            ]
            for row in workbook.history
        ],
        {"Content": 70, "Topic": 24, "Performance": 26},
        wrap_columns={"Content"},
    )

    monetization_headers = ["Metric", "Current", "Target", "Status", "Note"]
    _write_sheet(
        book.create_sheet("Monetization"),
        monetization_headers,
        [
            [row.metric, row.current, row.target, row.status, row.note]
            for row in workbook.monetization
        ],
        {"Metric": 38, "Current": 16, "Target": 16, "Status": 16, "Note": 60},
        wrap_columns={"Note"},
    )

    return book


def write(workbook: WeeklyWorkbook, directory: Path) -> Path:
    """Write the workbook and return where it landed."""
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / workbook.filename()
    build(workbook).save(path)
    return path


def summarise(workbook: WeeklyWorkbook) -> list[str]:
    """The lines printed after a run, so the week can be judged before opening it."""
    per_platform: dict[str, int] = {}
    per_topic: dict[str, int] = {}
    scores: list[float] = []
    needs_review = 0
    high_slop = 0

    for row in workbook.calendar:
        per_platform[row.platform] = per_platform.get(row.platform, 0) + 1
        per_topic[row.topic] = per_topic.get(row.topic, 0) + 1
        if row.authenticity is not None:
            scores.append(row.authenticity)
        if row.status == DraftStatus.REVIEW.value:
            needs_review += 1
        if row.slop_risk == "HIGH":
            high_slop += 1

    lines = [f"{len(workbook.calendar)} pieces scheduled"]
    for platform in (Platform.X.value, Platform.LINKEDIN.value):
        if platform in per_platform:
            lines.append(f"  {platform}: {per_platform[platform]}")
    if scores:
        lines.append(f"Average authenticity: {sum(scores) / len(scores):.1f}/10")
    lines.append(f"High slop risk: {high_slop}")
    lines.append(f"Needs review: {needs_review}")
    lines.append(f"Unused ideas kept: {len(workbook.ideas)}")

    top = sorted(per_topic.items(), key=lambda item: (-item[1], item[0]))[:8]
    if top:
        lines.append("Topics: " + ", ".join(f"{name} {count}" for name, count in top))
    return lines
